from typing import Union, Tuple

from django.utils import timezone
from openpyxl import load_workbook

from iiko.server import IikoServer
from .charts_services import FILE_ERRORS_MESSAGES
from .dataclasses.create_document_request import DocumentData
from .dataclasses.create_document_response import Response
from .models import Chain, Product, Department, Document


def excel_file_handler(file) -> Union[Tuple[DocumentData, dict], Tuple[None, dict]]:
    """
    Проверка и преобразование xlsx-файла в pydantic-класс
    с генерацией ошибок валидации
    """
    errors = {
        'empty_field': [],
        'wrong_type_of_filed': [],
        'no_product': [],
        'a_lot_of_products': [],
        'no_department': [],
        'a_lot_of_departments': [],
        'any_other': []
    }
    try:
        workbook = load_workbook(file)
        first_sheet = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(first_sheet)
        items = []
        for i, row in enumerate(worksheet.iter_rows(min_row=2), start=1):
            try:
                add_item = True

                try:
                    _ = iter(row)
                except TypeError:
                    errors['empty_field'].append(i + 1)
                    continue

                if len(row) < 4:
                    errors['empty_field'].append(i + 1)
                    continue

                if (
                    not row[0] or
                    not row[2] or
                    not row[3] or
                    row[0].value is None or
                    row[2].value is None or
                    row[3].value is None
                ):
                    errors['empty_field'].append(i + 1)
                    continue

                department_name = str(row[3].value)

                price = row[2].value
                if type(price) == str:
                    price = price.replace(',', '.')
                    price = price.replace(u'\xa0', '')
                    try:
                        price = float(price)
                    except Exception:
                        errors['wrong_type_of_filed'].append(i + 1)
                        continue
                else:
                    price = float(price)

                num = row[0].value
                if type(num) == int:
                    num = str(row[0].value)
                elif type(num) == float:
                    num = str(int(num))
                else:
                    num = str(num)

                product = Product.objects.filter(num=num)
                if product.count() == 0:
                    errors['no_product'].append(i + 1)
                    add_item = False
                if product.count() >= 2:
                    errors['a_lot_of_products'].append(i + 1)
                    add_item = False

                department = Department.objects.filter(name=department_name)
                if department.count() == 0:
                    errors['no_department'].append(i + 1)
                    add_item = False
                if department.count() >= 2:
                    errors['a_lot_of_departments'].append(i + 1)
                    add_item = False

                if add_item:
                    item = {
                        'num': i,
                        'product_id': str(product.first().uuid),
                        'price': price,
                        'department_id': str(department.first().uuid),
                        'dish_of_day': False,
                        'including': True,
                        'flyer_program': False,
                        'product_size_id': None
                    }
                    items.append(item)
            except Exception as e:
                errors['any_other'].append(e)
                continue

        document_data = {
            'date_incoming': str((timezone.localtime(timezone.now())).strftime('%Y-%m-%d')),
            'status': 'NEW',
            'delete_previous_menu': False,
            'items': items,
            'short_name': ''
        }
        return DocumentData(**document_data), errors

    except Exception as e:
        errors['any_other'].append(e)
        return None, errors


def create_file_errors_description(errors):
    """
    Генерация описания ошибок парсинга файла.
    """
    errors_description = []
    for error_type, string_numbers in errors.items():
        if len(string_numbers) != 0:
            errors_description.append(
                FILE_ERRORS_MESSAGES[error_type] + ', '.join([str(number) for number in string_numbers]))
    return errors_description


def create_documents_errors_description(response: Response) -> list[str]:
    """
    Генерация описания ошибок создания приказа.
    """
    errors_descriptions = []
    for error in response.errors:
        if 'Product size is not specified in position' in error.value:
            num = error.value.split('position ')[1].replace('.', '')
            for item in response.response.items:
                if item.num == int(num):
                    product = Product.objects.filter(uuid=item.uuid)
                    if product.count() != 0:
                        product_name = product.first().name
                        errors_descriptions.append(f'У блюда "{product_name}" не указан размер')
                        break
        elif 'No cooking place type for product' in error.value:
            uuid = error.value.split('product ')[1].split(' in position')[0]
            product = Product.objects.filter(uuid=uuid)
            if product.count() != 0:
                product_name = product.first().name
                errors_descriptions.append(f'У блюда "{product_name}" не указан тип места приготовления')
        elif 'Items is not specified' in error.value:
            errors_descriptions.append(f'Не добавлено ни одной позиции для формирования приказа')
        else:
            errors_descriptions.append(error.value)
    return errors_descriptions


def document_object_handler(obj: Document) -> Document:
    """
    Обработчик создания объекта Document.
    """
    obj.file_name = obj.file.name
    document, errors = excel_file_handler(obj.file)
    obj.file.delete()
    if document is None:
        obj.check_file = False
        obj.file_errors = errors['any_other']
        return obj
    errors_description = create_file_errors_description(errors)
    if errors_description:
        obj.check_file = False
        obj.file_errors = '\n'.join(errors_description)
    else:
        obj.check_file = True
    iiko_server = IikoServer(Chain.objects.first())
    response = iiko_server.document(document)
    if response.status_code == 200:
        response = Response.model_validate_json(response.text)
        if response.result == 'SUCCESS':
            obj.check_document = True
            obj.document_number = response.response.document_number
            return obj
        elif response.result == 'ERROR':
            errors_description = create_documents_errors_description(response)
            obj.check_document = False
            obj.document_errors = '\n'.join(errors_description)
            return obj
    else:
        obj.document_errors = response.text
        return obj
