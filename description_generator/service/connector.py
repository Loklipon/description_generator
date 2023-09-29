from datetime import datetime, timedelta

from openpyxl import load_workbook

from service.dataclasses.document import DocumentData
from service.models import Product, Department


def from_xlsx_file_data_to_server_create_data(file):
    """
    Преобразование xlsx-файла в pydantic-класс.
    """
    document_data = {
        'date_incoming': str((datetime.now() + timedelta(hours=3)).strftime('%Y-%m-%d')),
        'status': 'NEW',
        'items': [],
        'delete_previous_menu': False,
        'short_name': ''
    }
    workbook = load_workbook(file)
    first_sheet = workbook.get_sheet_names()[0]
    worksheet = workbook.get_sheet_by_name(first_sheet)
    for i, row in enumerate(worksheet.iter_rows(min_row=2), start=1):
        if row[0].value is None or row[2].value is None or row[3].value is None:
            continue
        product = Product.objects.filter(num=row[0].value)
        department = Department.objects.filter(name=row[3].value)
        if product.count() != 1 or department.count() != 1:
            return None
        item = {
            'num': i,
            'product_id': str(product.first().uuid),
            'price': row[2].value,
            'department_id': str(department.first().uuid),
            'dish_of_day': False,
            'including': True,
            'flyer_program': False,
            'product_size_id': None
        }
        document_data['items'].append(item)
    return DocumentData(**document_data)
